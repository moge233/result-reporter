#! python3

import csv
from datetime import datetime
from enum import Enum
from math import floor, nan
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from pydrf.textchart import Header, RaceData, StarterPerformanceData, RecordType, CourseCodes

from .chart import Chart
from .coursetype import CourseType
from .report import BrohamerReport, ShakeUpReport, DEFAULT_MAXIMUM_SPRINT_DISTANCE


TAKEOUT_PCT: float = 0.2


class DistanceKey(Enum):
    SPRINT = 0
    ROUTE = 1


def parse_chart(path: str) -> Chart | None:
    header: Header | None = None
    race_data: list[RaceData] = []
    starters_performance_data: list[StarterPerformanceData] = []
    try:
        with open(path) as chart_file:
            reader = csv.reader(chart_file.readlines())
            for line in reader:
                if line[0] == RecordType.HEADER:
                    header = Header.create(line)
                elif line[0] == RecordType.RACE:
                    race_data.append(RaceData.create(line))
                elif line[0] == RecordType.STARTER:
                    starters_performance_data.append(StarterPerformanceData.create(line))
                elif line[0] == RecordType.EXOTIC_WAGERING:
                    pass
                elif line[0] == RecordType.ATTENDANCE:
                    pass
                elif line[0] == RecordType.COMMENT:
                    pass
                elif line[0] == RecordType.FOOTNOTE:
                    pass
        if header and race_data and starters_performance_data:
            return Chart(
                header,
                race_data,
                starters_performance_data
            )
        return None
    except FileNotFoundError as e:
        print(f'[{e}]: could not find file {path}')
        return None


def get_charts(path: str, track_code: str) -> list[Chart]:
    charts: list[Chart] = []
    for dir in os.listdir(path):
        dir_path = os.path.join(path, dir)
        for chart_path in os.listdir(dir_path):
            if chart_path[:len(track_code)] == track_code and \
                    chart_path[len(track_code)].isdigit():
                chart_path = os.path.join(dir_path, chart_path)
                chart: Chart | None = parse_chart(chart_path)
                if chart:
                    charts.append(chart)
    return charts


def get_surfaces(charts: list[Chart]) -> list[str]:
    surfaces: list[str] = []
    for chart in charts:
        for race in chart.races:
            if race.data.course_type not in surfaces:
                surfaces.append(race.data.course_type)
    return sorted(surfaces)


def get_course_types(charts: list[Chart]) -> list[CourseType]:
    course_types: list[CourseType] = []
    for chart in charts:
        for race in chart.races:
            course_type: CourseType = CourseType.parse_course_type(race.data.course_type)
            if course_type not in course_types:
                course_types.append(course_type)
    return course_types


def shakeup_report_distance_key_match(distance_key: DistanceKey, report: ShakeUpReport) -> bool:
    return (((distance_key == DistanceKey.SPRINT) and (report.distance <= DEFAULT_MAXIMUM_SPRINT_DISTANCE)) or
            ((distance_key == DistanceKey.ROUTE) and (report.distance > DEFAULT_MAXIMUM_SPRINT_DISTANCE)))


def distance_key_match(distance_key: DistanceKey, distance: float) -> bool:
    return (((distance_key == DistanceKey.SPRINT) and (distance <= DEFAULT_MAXIMUM_SPRINT_DISTANCE)) or
            ((distance_key == DistanceKey.ROUTE) and (distance > DEFAULT_MAXIMUM_SPRINT_DISTANCE)))


def course_to_str(course: str) -> str:
    if course == CourseCodes.ALL_WEATHER_TRACK.value:
        return 'Tapeta'
    elif course == CourseCodes.DIRT.value:
        return 'Dirt'
    elif course == CourseCodes.HURDLE.value:
        return 'Hurdle'
    elif course == CourseCodes.INNER_TRACK.value:
        return 'Inner Track'
    elif course == CourseCodes.INNER_TURF.value:
        return 'Inner Turf'
    elif course == CourseCodes.OUTER_TURF.value:
        return 'Outer Turf'
    elif course == CourseCodes.TURF.value:
        return 'Turf'
    return ''


def course_to_excel_color(course: str) -> str:
    if course == CourseCodes.ALL_WEATHER_TRACK.value:
        return 'FF8C00'
    elif course == CourseCodes.DIRT.value:
        return '000000'
    elif course == CourseCodes.HURDLE.value:
        return '000000'
    elif course == CourseCodes.INNER_TRACK.value:
        return '000000'
    elif course == CourseCodes.INNER_TURF.value:
        return '154734'
    elif course == CourseCodes.OUTER_TURF.value:
        return '154734'
    elif course == CourseCodes.TURF.value:
        return "178F17"
    return ''


class OddsLevel(Enum):
    LEVEL_ONE = 1
    LEVEL_TWO = 2
    LEVEL_THREE = 3
    LEVEL_FOUR = 4


def get_odds_score(odds: float) -> float:
    if odds <= 4.0:
        return 1.0
    elif odds <= 8.0:
        return 2.0
    elif odds <= 12.0:
        return 3.0
    else:
        return 4.0


def get_expected_value_from_odds(odds: float) -> float:
    our_odds = odds / 100.0
    if our_odds <= 1:
        ret = (our_odds / (our_odds + 1)) * (1 - TAKEOUT_PCT)
    else:
        ret = (1 / (our_odds + 1)) * (1 - TAKEOUT_PCT)
    assert ret < 1
    return ret


def get_post_position_bias_comment(surface: str, distance_key: DistanceKey, chart: Chart) -> str:
    ret = ''
    total_count: int = 0
    w_count13: int = 0
    w_count46: int = 0
    w_count79: int = 0
    w_count1012: int = 0
    w_total_odds13: float = 0
    w_total_odds46: float = 0
    w_total_odds79: float = 0
    w_total_odds1012: float = 0
    w_average_odds13: float = 0
    w_average_odds46: float = 0
    w_average_odds79: float = 0
    w_average_odds1012: float = 0
    for race in chart.races:
        if race.data.breed_indicator != 'TB':
            continue
        winner: StarterPerformanceData = race.starters[0]
        w_post_position: int = winner.post_position
        w_odds: float = winner.odds / 100.0
        if distance_key_match(distance_key, race.data.distance / 100) \
                and (race.data.course_type == surface):
            total_count += 1
            # Winner post positions
            if w_post_position == 1 or w_post_position <= 3:
                w_count13 += 1
                w_total_odds13 += get_odds_score(w_odds)
            elif w_post_position == 4 or w_post_position <= 6:
                w_count46 += 1
                w_total_odds46 += get_odds_score(w_odds)
            elif w_post_position == 7 or w_post_position <= 9:
                w_count79 += 1
                w_total_odds79 += get_odds_score(w_odds)
            elif w_post_position == 10 or w_post_position <= 12:
                w_count1012 += 1
                w_total_odds1012 += get_odds_score(w_odds)
    if total_count == 2:
        return '2r'
    elif total_count == 1:
        return '1r'
    elif total_count == 0:
        return 'nr'
    else:
        w_average_odds13 = w_total_odds13 / w_count13 if w_count13 else 0.0
        w_average_odds46 = w_total_odds46 / w_count46 if w_count46 else 0.0
        w_average_odds79 = w_total_odds79 / w_count79 if w_count79 else 0.0
        w_average_odds1012 = w_total_odds1012 / w_count1012 if w_count1012 else 0.0
        sorted_scores: list[float] = sorted(
            [w_average_odds13, w_average_odds46, w_average_odds79, w_average_odds1012],
            reverse=True
        )
        high_score: float = sorted_scores[0]
        second_score: float = sorted_scores[1]
        percent_diff: float = round((high_score - second_score) / second_score, 2) if second_score else 2.0
        if percent_diff > 1.0:
            if high_score == w_average_odds13:
                # Inside
                ret += 'In'
            elif high_score != w_average_odds13:
                ret += 'Br'
    return ret


def get_running_style_bias_comment(surface: str, distance_key: DistanceKey, chart: Chart) -> str:
    ret = ''
    total_count: int = 0
    w_count_e: int = 0
    w_count_p: int = 0
    w_count_s: int = 0
    w_total_odds_e: float = 0
    w_total_odds_p: float = 0
    w_total_odds_s: float = 0
    w_average_odds_e: float = 0
    w_average_odds_p: float = 0
    w_average_odds_s: float = 0
    for race in chart.races:
        if race.data.breed_indicator != 'TB':
            continue
        winner: StarterPerformanceData = race.starters[0]
        if distance_key is DistanceKey.SPRINT:
            w_bl1: float = winner.length_behind_at_poc1 / 100
        else:
            w_bl1: float = winner.length_behind_at_poc2 / 100
        w_odds: float = winner.odds / 100.0
        if distance_key_match(distance_key, race.data.distance / 100) \
                and (race.data.course_type == surface):
            total_count += 1
            if w_bl1 < 1:
                w_count_e += 1
                w_total_odds_e += get_odds_score(w_odds)
            elif w_bl1 >= 1 and w_bl1 <= 5:
                w_count_p += 1
                w_total_odds_p += get_odds_score(w_odds)
            else:
                w_count_s += 1
                w_total_odds_s += get_odds_score(w_odds)
    if total_count == 2:
        return '2r'
    elif total_count == 1:
        return '1r'
    elif total_count == 0:
        return 'nr'
    else:
        w_average_odds_e = w_total_odds_e / w_count_e if w_count_e else 0.0
        w_average_odds_p = w_total_odds_p / w_count_p if w_count_p else 0.0
        w_average_odds_s = w_total_odds_s / w_count_s if w_count_s else 0.0
        sorted_scores: list[float] = sorted([w_average_odds_e, w_average_odds_p, w_average_odds_s], reverse=True)
        high_score: float = sorted_scores[0]
        second_score: float = sorted_scores[1]
        percent_diff: float = round((high_score - second_score) / second_score, 2) if second_score else 1.9
        if percent_diff > 1.0:
            if high_score == w_average_odds_e:
                # Speed
                ret += 'Sp'
            elif high_score == w_average_odds_p:
                ret += 'St'
            elif high_score == w_average_odds_s:
                ret += 'Cl'
    return ret


def get_daily_comment(surface: str, distance_key: DistanceKey, chart: Chart) -> str:
    ret: str = ''
    pp_bias_comment: str = get_post_position_bias_comment(surface, distance_key, chart)
    rs_bias_comment: str = get_running_style_bias_comment(surface, distance_key, chart)
    if pp_bias_comment in ('2r', '1r', 'nr'):
        return pp_bias_comment
    else:
        if pp_bias_comment != '':
            ret += pp_bias_comment
        if rs_bias_comment != '':
            if pp_bias_comment != '':
                ret += ','
            ret += rs_bias_comment
    if ret == '':
        ret = '-'
    return ret


def decimal_to_fifths(frac: float) -> float:
    '''
    Take a regular decimal number and return a new float
    that represents the argument as a decimal where the decimal
    place represents fifths (instead of tenths)
    '''
    if frac is nan:
        return frac
    t1_tenths = floor(frac * 5) / 5
    integer_part = floor(t1_tenths)
    fractional_part = round(t1_tenths - integer_part, 1) / 2
    return integer_part + fractional_part


def get_shakeup_daily_maximums(surface: str, distance_key: DistanceKey, reports: list[ShakeUpReport]) -> \
        tuple[float, float, float]:
    maximum1: float = 0
    maximum2: float = 0
    maximum3: float = 0
    for report in reports:
        if report.surface == surface and shakeup_report_distance_key_match(distance_key, report):
            if report.fr1 > maximum1:
                maximum1 = report.fr1
            if report.fr2 > maximum2:
                maximum2 = report.fr2
            if report.fr3 > maximum3:
                maximum3 = report.fr3
    if maximum1 == 0:
        maximum1 = nan
    if maximum2 == 0:
        maximum2 = nan
    if maximum3 == 0:
        maximum3 = nan
    return (maximum1, maximum2, maximum3)


def get_shakeup_daily_minimums(surface: str, distance_key: DistanceKey, reports: list[ShakeUpReport]) -> \
        tuple[float, float, float]:
    minimum1: float = 1000
    minimum2: float = 1000
    minimum3: float = 1000
    for report in reports:
        if report.surface == surface and shakeup_report_distance_key_match(distance_key, report):
            if report.fr1 < minimum1:
                minimum1 = report.fr1
            if report.fr2 < minimum2:
                minimum2 = report.fr2
            if report.fr3 < minimum3:
                minimum3 = report.fr3
    if minimum1 == 1000:
        minimum1 = nan
    if minimum2 == 1000:
        minimum2 = nan
    if minimum3 == 1000:
        minimum3 = nan
    return (minimum1, minimum2, minimum3)


def get_shakeup_reports(chart: Chart) -> list[ShakeUpReport]:
    chart_reports: list[ShakeUpReport] = []
    for race in chart.races:
        if race.data.breed_indicator != 'TB':
            continue
        winner: StarterPerformanceData = race.starters[0]
        assert winner.official_finish == 1
        report = ShakeUpReport(
            key=f'{chart.race_date}{race.data.race_number:02d}',
            cls=race.data.race_type,
            claiming_price=race.data.maximum_claiming_price,
            purse=race.data.purse,
            surface=race.data.course_type,
            distance=race.data.distance,
            post_position=winner.post_position,
            bl1=winner.length_behind_at_poc1,
            bl2=winner.length_behind_at_poc2,
            bl3=winner.length_behind_at_poc3,
            blf=winner.length_behind_at_finish,
            fr1=race.data.fraction1,
            fr2=race.data.fraction2,
            fr3=race.data.fraction3,
            finish=race.data.final_time
        )
        chart_reports.append(report)
    return chart_reports


def create_hearts_guide(charts: list[Chart], path: str) -> None:
    '''
    Header:
    (empty)  |  (Surface #1) Sprints  |  (Surface #1) Routes  |  (Surface #2) Sprints  |  (Surface #2) Routes  |  etc...
    (1 cell) |  (4 cells)             |  (4 cells)            |  (4 cells)             |  (4 cells)            |  etc...

    Row 1: Header
    Row 2: Race date 1 (minimums)
    Row 3: Race date 1 (maximums)
    Row 4: Race date 2 (minimums)
    Row 5: Race date 2 (maximums)
    etc...

    Column A: Dates
    Column B: 2F Mins and Maxes
    Column C: 4F Mins and Maxes
    Column D: 6F Mins and Maxes
    Column E: Comment
    Column F: Repeat, without the date, if necessary
    '''
    if os.path.exists(path):
        raise FileExistsError(f'{path} already exists')
    wb: Workbook = Workbook()
    ws: Worksheet = wb.create_sheet('Sheet1')
    wb.active = ws
    # Some constants
    distance_keys: list[str] = ['Sprints', 'Routes']
    surfaces: list[str] = get_surfaces(charts)
    count_surfaces = len(surfaces)
    center_alignment = Alignment(horizontal="center", vertical="center")
    font = Font(name='Aptos Narrow', size=14, bold=True, color="000000")
    # TODO: Get the different colors for each surface (dirt - black, turf - green, tapeta - orange, etc.)
    # Add headers
    row_idx: int = 1
    column_idx: str = 'B'  # skip column A in the header
    cell_idx: str = f'{column_idx}{row_idx}'
    for surface in surfaces:
        for distance_key in distance_keys:
            header_value = f'{course_to_str(surface)} {distance_key}'
            ws[cell_idx] = header_value
            end_column_idx: str = chr(ord(column_idx) + 3)
            end_cell_idx = f'{end_column_idx}{row_idx}'
            ws.merge_cells(f'{cell_idx}:{end_cell_idx}')
            ws[cell_idx].alignment = center_alignment
            color = course_to_excel_color(surface)
            font = Font(name='Aptos Narrow', size=14, bold=True, color=color)
            ws[cell_idx].font = font
            column_idx = chr(ord(end_column_idx) + 1)
            # column_idx = increment_column_index(column_idx)
            cell_idx: str = f'{column_idx}{row_idx}'
    # Add rows of data
    row_idx = 2
    column_idx = 'A'
    cell_idx: str = f'{column_idx}{row_idx}'
    for chart in charts:
        reports: list[ShakeUpReport] = get_shakeup_reports(chart)
        row1: list[str | float] = []
        row2: list[str | float] = []
        date: datetime = datetime.strptime(chart.race_date, '%Y%m%d')
        row1.append(date.strftime('%m/%d'))
        row2.append('')
        for surface in surfaces:
            for distance_key in (DistanceKey.SPRINT, DistanceKey.ROUTE):
                minimums: tuple[float, float, float] = get_shakeup_daily_minimums(surface, distance_key, reports)
                maximums: tuple[float, float, float] = get_shakeup_daily_maximums(surface, distance_key, reports)
                for minimum in minimums:
                    row1.append(decimal_to_fifths(minimum))
                for maximum in maximums:
                    row2.append(decimal_to_fifths(maximum))
                # TODO: Generate comment for each surface and distance key combination
                row1.append(get_daily_comment(surface, distance_key, chart))
                row2.append('')
        row1 = ['-' if val is nan else val for val in row1]
        row2 = ['-' if val is nan else val for val in row2]
        ws.append(row1)
        ws.append(row2)
        # Merge the date cells
        end_row_idx: int = row_idx + 1
        end_cell_idx: str = f'{column_idx}{end_row_idx}'
        ws.merge_cells(f'{cell_idx}:{end_cell_idx}')
        font = Font(name='Aptos Narrow', size=14, bold=True, color='000000')
        ws[cell_idx].alignment = center_alignment
        ws[cell_idx].font = font
        # Iterate over the min. and max. fraction cells and comment cell of each row and set the font
        for i in range(2 * count_surfaces):
            color = course_to_excel_color(surfaces[i // 2])
            font = Font(name='Aptos Narrow', size=14, bold=True, color=color)
            for __ in range(1, 5):
                column_idx = chr(ord(column_idx) + 1)
                cell_idx = f'{column_idx}{row_idx}'
                ws[cell_idx].alignment = center_alignment
                ws[cell_idx].font = font
                cell_idx = f'{column_idx}{row_idx + 1}'
                ws[cell_idx].alignment = center_alignment
                ws[cell_idx].font = font
        # Reset to column A; the next row index is +2
        row_idx += 2
        column_idx = 'A'
        cell_idx: str = f'{column_idx}{row_idx}'
    wb.save(path)


def get_brohamer_daily_maximums(surface: CourseType, distance_key: DistanceKey, reports: list[BrohamerReport]) -> \
        tuple[float, float, float]:
    maximum1: float = 0
    maximum2: float = 0
    maximum3: float = 0
    for report in reports:
        if CourseType.parse_course_type(report.course) == surface and distance_key_match(distance_key, report.distance):
            if report.fr1 > maximum1:
                maximum1 = report.fr1
            if report.fr2 > maximum2:
                maximum2 = report.fr2
            if report.fr3 > maximum3:
                maximum3 = report.fr3
    if maximum1 == 0:
        maximum1 = nan
    if maximum2 == 0:
        maximum2 = nan
    if maximum3 == 0:
        maximum3 = nan
    return (maximum1, maximum2, maximum3)


def get_brohamer_daily_minimums(surface: CourseType, distance_key: DistanceKey, reports: list[BrohamerReport]) -> \
        tuple[float, float, float]:
    minimum1: float = 10000
    minimum2: float = 10000
    minimum3: float = 10000
    for report in reports:
        if CourseType.parse_course_type(report.course) == surface and distance_key_match(distance_key, report.distance):
            if report.fr1 < minimum1:
                minimum1 = report.fr1
            if report.fr2 < minimum2:
                minimum2 = report.fr2
            if report.fr3 < minimum3:
                minimum3 = report.fr3
    if minimum1 == 10000:
        minimum1 = nan
    if minimum2 == 10000:
        minimum2 = nan
    if minimum3 == 10000:
        minimum3 = nan
    return (minimum1, minimum2, minimum3)


def get_brohamer_comment(surface: CourseType, distance_key: DistanceKey, reports: list[BrohamerReport]) -> str:
    ret = ''
    return ret


def get_brohamer_reports(chart: Chart) -> list[BrohamerReport]:
    chart_reports: list[BrohamerReport] = []
    for race in chart.races:
        if race.data.breed_indicator != 'TB':
            continue
        winner: StarterPerformanceData = race.starters[0]
        assert winner.official_finish == 1
        c1: float = race.data.fraction1 if race.data.distance / 100 <= DEFAULT_MAXIMUM_SPRINT_DISTANCE else \
            race.data.fraction2
        c2: float = race.data.fraction2 if race.data.distance / 100 <= DEFAULT_MAXIMUM_SPRINT_DISTANCE else \
            race.data.fraction3
        bl1: float = winner.length_behind_at_poc1 if race.data.distance / 100 <= DEFAULT_MAXIMUM_SPRINT_DISTANCE else \
            winner.length_behind_at_poc2
        bl2: float = winner.length_behind_at_poc2 if race.data.distance / 100 <= DEFAULT_MAXIMUM_SPRINT_DISTANCE else \
            winner.length_behind_at_poc3
        report: BrohamerReport = BrohamerReport(
            key=f'{chart.race_date}{race.data.race_number:02d}',
            cls=race.data.race_type,
            sex=race.data.sex_restriction,
            age=race.data.age_restriction,
            claiming_price=race.data.maximum_claiming_price,
            purse=race.data.purse,
            race=race.data.race_number,
            surface=race.data.surface,
            course=race.data.course_type,
            distance=race.data.distance,
            number=race.data.number_of_horses,
            post=winner.post_position,
            bl1=bl1,
            bl2=bl2,
            c1=c1,
            c2=c2,
            fc=race.data.final_time
        )
        chart_reports.append(report)
    return chart_reports


def create_brohamer_guide(charts: list[Chart], path: str) -> None:
    '''
    Header:
    (empty)  |  (Surface #1) Sprints  |  (Surface #1) Routes  |  (Surface #2) Sprints  |  (Surface #2) Routes  |  etc...
    (1 cell) |  (4 cells)             |  (4 cells)            |  (4 cells)             |  (4 cells)            |  etc...

    Row 1: Header
    Row 2: Race date 1 (minimums)
    Row 3: Race date 1 (maximums)
    Row 4: Race date 2 (minimums)
    Row 5: Race date 2 (maximums)
    etc...

    Column A: Dates
    Column B: FR1 Mins and Maxes
    Column C: FR2 Mins and Maxes
    Column D: FR3 Mins and Maxes
    Column E: Comment
    Column F: Repeat, without the date, if necessary
    '''
    if os.path.exists(path):
        raise FileExistsError(f'{path} already exists')
    wb: Workbook = Workbook()
    ws: Worksheet = wb.create_sheet('Sheet1')
    wb.active = ws
    # Some constants
    distance_keys: list[str] = ['Sprints', 'Routes']
    surfaces: list[str] = get_surfaces(charts)
    count_surfaces = len(surfaces)
    center_alignment = Alignment(horizontal="center", vertical="center")
    font = Font(name='Aptos Narrow', size=14, bold=True, color="000000")
    # Add headers
    row_idx: int = 1
    column_idx: int = 2  # skip column A in the header
    cell_idx = ws.cell(row=row_idx, column=column_idx)
    for surface in surfaces:
        for distance_key in distance_keys:
            header_value = f'{course_to_str(surface)} {distance_key}'
            ws[cell_idx.coordinate] = header_value
            end_column_idx: int = column_idx + 3
            end_cell_idx = ws.cell(row=row_idx, column=end_column_idx)
            ws.merge_cells(f'{cell_idx.coordinate}:{end_cell_idx.coordinate}')
            ws[cell_idx.coordinate].alignment = center_alignment
            color = course_to_excel_color(surface)
            font = Font(name='Aptos Narrow', size=14, bold=True, color=color)
            ws[cell_idx.coordinate].font = font
            column_idx = end_column_idx + 1
            cell_idx = ws.cell(row=row_idx, column=column_idx)
    # Add data
    row_idx = 2
    column_idx = 1
    for chart in charts:
        reports: list[BrohamerReport] = get_brohamer_reports(chart)
        row1: list[str | float] = []
        row2: list[str | float] = []
        date: datetime = datetime.strptime(chart.race_date, '%Y%m%d')
        row1.append(date.strftime('%m/%d'))
        row2.append('')
        for surface in surfaces:
            for distance_key in (DistanceKey.SPRINT, DistanceKey.ROUTE):
                course_type: CourseType = CourseType.parse_course_type(surface)
                minimums: tuple[float, float, float] = get_brohamer_daily_minimums(course_type, distance_key, reports)
                maximums: tuple[float, float, float] = get_brohamer_daily_maximums(course_type, distance_key, reports)
                for minimum in minimums:
                    row1.append(round(minimum, 1))
                for maximum in maximums:
                    row2.append(round(maximum, 1))
                row1.append(get_daily_comment(surface, distance_key, chart))
                row2.append('')
        row1 = ['-' if val is nan else val for val in row1]
        row2 = ['-' if val is nan else val for val in row2]
        ws.append(row1)
        ws.append(row2)
        # Merge the date cells
        cell_idx = ws.cell(row=row_idx, column=column_idx)
        end_row_idx: int = row_idx + 1
        end_cell_idx = ws.cell(row=end_row_idx, column=column_idx)
        ws.merge_cells(f'{cell_idx.coordinate}:{end_cell_idx.coordinate}')
        font = Font(name='Aptos Narrow', size=14, bold=True, color='000000')
        ws[cell_idx.coordinate].alignment = center_alignment
        ws[cell_idx.coordinate].font = font
        # Iterate over the min. and max. fraction cells and comment cell of each row and set the font
        for i in range(2 * count_surfaces):
            color = course_to_excel_color(surfaces[i // 2])
            font = Font(name='Aptos Narrow', size=14, bold=True, color=color)
            for __ in range(1, 5):
                column_idx += 1
                cell_idx = ws.cell(row=row_idx, column=column_idx)
                ws[cell_idx.coordinate].alignment = center_alignment
                ws[cell_idx.coordinate].font = font
                cell_idx = ws.cell(row=row_idx + 1, column=column_idx)
                ws[cell_idx.coordinate].alignment = center_alignment
                ws[cell_idx.coordinate].font = font
        # Reset to column A; the next row index is current row + 2
        row_idx += 2
        column_idx = 1
    wb.save(path)


def create_brohamer_day_report(chart: Chart, path: str) -> None:
    reports: list[BrohamerReport] = get_brohamer_reports(chart)
    for report in reports:
        print(report)
    return
